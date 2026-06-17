"""
NIFTY Quant Lab - Daily Report Generator
==========================================
Generates end-of-day market reports in three formats:
  - HTML  (for Telegram/web preview)
  - Excel (for quantitative analysis)
  - PDF   (for archival / sharing)

Report sections:
  1. Market Overview (NIFTY, BANKNIFTY, VIX)
  2. Top BUY Signals (today's scanner output)
  3. Sector Performance Heatmap
  4. India VIX + PCR Context
  5. Top Movers (gainers/losers)
  6. Key Events Tomorrow

Architecture:
- PyPortfolioOpt risk_models pattern: clean computation → clean output
- vectorbt Portfolio stats pattern: one method per metric category
- gs-quant DataGrid pattern: structured multi-section report assembly
"""

from __future__ import annotations

import logging
import os
from datetime import date, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from nifty_quant_lab.config.settings import settings, NIFTY50_SYMBOLS
from nifty_quant_lab.utils.logger import get_logger

logger = get_logger("report_generator")


class DailyReportGenerator:
    """
    Generates structured market reports with data from the DB.

    Output formats: HTML → Excel → PDF (produced in that order;
    each can be sent independently via Telegram).
    """

    def __init__(self):
        self.output_dir = settings.reports_dir
        self.today = date.today()

    async def generate_and_send(self) -> Dict[str, Optional[str]]:
        """
        Full pipeline: gather data → build report → save → send via Telegram.

        Returns:
            Paths to generated files: {html, excel, pdf}
        """
        logger.info(f"Generating daily report for {self.today}...")

        data = await self._gather_data()
        if not data:
            logger.warning("No data gathered — aborting report")
            return {}

        paths = {}
        paths["html"] = await self._generate_html(data)
        paths["excel"] = await self._generate_excel(data)
        paths["pdf"] = self._html_to_pdf(paths["html"]) if paths["html"] else None

        await self._persist_report_metadata(data, paths)
        await self._send_report(data, paths)

        logger.info(f"Daily report complete: {paths}")
        return paths

    # ──────────────────────────────────────────────────────────
    # DATA GATHERING
    # ──────────────────────────────────────────────────────────

    async def _gather_data(self) -> Optional[Dict]:
        """Collect all data needed for the report from DB and live sources."""
        try:
            from nifty_quant_lab.database.connection import get_async_session
            from nifty_quant_lab.database.models import (
                HistoricalPrice, ScannerResult, Symbol, SignalType
            )
            from sqlalchemy import select, desc, func

            data = {"date": self.today, "sections": {}}

            async with get_async_session() as db:
                # ── Market Overview: NIFTY, BANKNIFTY, VIX
                overview = {}
                for index_sym in ["NIFTY50", "BANKNIFTY", "INDIA_VIX"]:
                    sym_res = await db.execute(
                        select(Symbol.id).where(
                            Symbol.symbol == index_sym, Symbol.exchange == "NSE"
                        )
                    )
                    sid = sym_res.scalar_one_or_none()
                    if not sid:
                        continue
                    price_res = await db.execute(
                        select(HistoricalPrice)
                        .where(HistoricalPrice.symbol_id == sid)
                        .order_by(desc(HistoricalPrice.date))
                        .limit(2)
                    )
                    rows = price_res.scalars().all()
                    if len(rows) >= 2:
                        latest, prev = rows[0], rows[1]
                        close = float(latest.close)
                        prev_close = float(prev.close)
                        change = close - prev_close
                        change_pct = change / prev_close * 100
                        overview[index_sym] = {
                            "close": close,
                            "open": float(latest.open),
                            "high": float(latest.high),
                            "low": float(latest.low),
                            "change": round(change, 2),
                            "change_pct": round(change_pct, 2),
                            "volume": int(latest.volume),
                        }
                data["sections"]["market_overview"] = overview

                # ── Top Scanner Signals
                scan_res = await db.execute(
                    select(ScannerResult, Symbol.symbol, Symbol.sector)
                    .join(Symbol, ScannerResult.symbol_id == Symbol.id)
                    .where(ScannerResult.scan_date == self.today)
                    .order_by(desc(ScannerResult.score))
                    .limit(20)
                )
                scan_rows = scan_res.all()
                buy_signals = [
                    {
                        "symbol": sym,
                        "sector": sec or "N/A",
                        "signal": r.signal.value,
                        "score": r.score,
                        "close": r.close_price,
                        "rsi": r.rsi,
                        "conditions": sum([
                            r.ema20_above_ema50, r.rsi_above_55, r.macd_bullish_cross,
                            r.price_above_supertrend, r.volume_above_avg, r.week52_breakout
                        ]),
                    }
                    for r, sym, sec in scan_rows
                    if r.signal == SignalType.BUY
                ]
                data["sections"]["buy_signals"] = buy_signals[:10]
                data["sections"]["total_buy"] = len(buy_signals)
                data["sections"]["total_scanned"] = len(scan_rows)

                # ── Sector Performance (last 5 days)
                since = self.today - timedelta(days=7)
                sector_data = {}
                for sym in list(NIFTY50_SYMBOLS)[:50]:
                    sym_res2 = await db.execute(
                        select(Symbol.id, Symbol.sector)
                        .where(Symbol.symbol == sym, Symbol.exchange == "NSE")
                    )
                    sym_row = sym_res2.first()
                    if not sym_row or not sym_row[1]:
                        continue
                    sid2, sector = sym_row
                    pr = await db.execute(
                        select(HistoricalPrice.close)
                        .where(
                            HistoricalPrice.symbol_id == sid2,
                            HistoricalPrice.date >= since,
                        )
                        .order_by(HistoricalPrice.date.asc())
                        .limit(6)
                    )
                    closes = [float(r[0]) for r in pr.all()]
                    if len(closes) >= 2:
                        perf = (closes[-1] - closes[0]) / closes[0] * 100
                        if sector not in sector_data:
                            sector_data[sector] = []
                        sector_data[sector].append(perf)

                sector_avg = {
                    sec: round(sum(perfs) / len(perfs), 2)
                    for sec, perfs in sector_data.items()
                    if perfs
                }
                data["sections"]["sector_performance"] = dict(
                    sorted(sector_avg.items(), key=lambda x: x[1], reverse=True)
                )

            return data

        except Exception as e:
            logger.error(f"Data gathering failed: {e}", exc_info=True)
            return None

    # ──────────────────────────────────────────────────────────
    # HTML REPORT
    # ──────────────────────────────────────────────────────────

    async def _generate_html(self, data: Dict) -> Optional[str]:
        """Generate styled HTML report."""
        try:
            html_path = self.output_dir / f"report_{self.today}.html"
            overview = data["sections"].get("market_overview", {})
            buy_signals = data["sections"].get("buy_signals", [])
            sectors = data["sections"].get("sector_performance", {})

            def _arrow(chg: float) -> str:
                return "▲" if chg >= 0 else "▼"

            def _color(chg: float) -> str:
                return "#00c853" if chg >= 0 else "#d50000"

            # Market cards
            mkt_cards = ""
            for sym, d in overview.items():
                label = {"NIFTY50": "NIFTY 50", "BANKNIFTY": "BANK NIFTY", "INDIA_VIX": "India VIX"}.get(sym, sym)
                clr = _color(d["change_pct"])
                arr = _arrow(d["change_pct"])
                mkt_cards += f"""
                <div class="card">
                  <h3>{label}</h3>
                  <div class="price">₹{d['close']:,.2f}</div>
                  <div class="change" style="color:{clr}">{arr} {d['change']:+,.2f} ({d['change_pct']:+.2f}%)</div>
                  <div class="meta">H: ₹{d['high']:,.2f} | L: ₹{d['low']:,.2f}</div>
                </div>"""

            # Signal rows
            signal_rows = ""
            for i, s in enumerate(buy_signals, 1):
                clr = "#aaa"
                signal_rows += f"""
                <tr>
                  <td>{i}</td>
                  <td><strong>{s['symbol']}</strong></td>
                  <td>{s.get('sector','')}</td>
                  <td><span class="badge-buy">BUY</span></td>
                  <td><strong>{s['score']:.0f}</strong></td>
                  <td>{s['conditions']}/6</td>
                  <td>₹{s['close']:,.2f}</td>
                  <td>{f"{s['rsi']:.1f}" if s['rsi'] else 'N/A'}</td>
                </tr>"""

            # Sector bars
            sector_rows = ""
            for sec, perf in list(sectors.items())[:12]:
                clr = _color(perf)
                bar_w = min(abs(perf) * 10, 100)
                direction = "right" if perf >= 0 else "left"
                sector_rows += f"""
                <div class="sector-row">
                  <span class="sector-name">{sec}</span>
                  <div class="bar-container">
                    <div class="bar" style="width:{bar_w:.0f}px;background:{clr};float:{direction}"></div>
                  </div>
                  <span class="sector-pct" style="color:{clr}">{perf:+.2f}%</span>
                </div>"""

            html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>NIFTY Quant Lab — Daily Report {self.today}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #0d1117; color: #e6edf3; padding: 24px; }}
  h1 {{ font-size: 24px; color: #58a6ff; margin-bottom: 4px; }}
  h2 {{ font-size: 16px; color: #8b949e; margin: 24px 0 12px; text-transform: uppercase; letter-spacing: 1px; border-bottom: 1px solid #21262d; padding-bottom: 6px; }}
  .subtitle {{ color: #8b949e; font-size: 13px; margin-bottom: 24px; }}
  .cards {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 8px; }}
  .card {{ background: #161b22; border: 1px solid #21262d; border-radius: 8px; padding: 18px 24px; min-width: 200px; }}
  .card h3 {{ color: #8b949e; font-size: 12px; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 8px; }}
  .price {{ font-size: 26px; font-weight: 700; color: #e6edf3; }}
  .change {{ font-size: 14px; margin-top: 4px; font-weight: 600; }}
  .meta {{ font-size: 12px; color: #8b949e; margin-top: 6px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ background: #161b22; color: #8b949e; font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px; padding: 10px 12px; text-align: left; }}
  td {{ padding: 10px 12px; border-bottom: 1px solid #21262d; }}
  tr:hover {{ background: #161b22; }}
  .badge-buy {{ background: #1a4731; color: #3fb950; padding: 2px 8px; border-radius: 12px; font-size: 11px; font-weight: 700; }}
  .sector-row {{ display: flex; align-items: center; gap: 12px; margin: 6px 0; }}
  .sector-name {{ width: 160px; font-size: 12px; color: #8b949e; }}
  .bar-container {{ width: 200px; height: 14px; background: #21262d; border-radius: 3px; overflow: hidden; }}
  .bar {{ height: 100%; border-radius: 3px; transition: width 0.3s; }}
  .sector-pct {{ font-size: 13px; font-weight: 600; min-width: 60px; }}
  .footer {{ margin-top: 40px; padding-top: 16px; border-top: 1px solid #21262d; color: #8b949e; font-size: 12px; }}
  .stat-row {{ display: flex; gap: 24px; margin-bottom: 16px; }}
  .stat {{ background: #161b22; border: 1px solid #21262d; border-radius: 6px; padding: 12px 18px; }}
  .stat-val {{ font-size: 22px; font-weight: 700; color: #58a6ff; }}
  .stat-lbl {{ font-size: 11px; color: #8b949e; margin-top: 2px; }}
</style>
</head>
<body>
<h1>📊 NIFTY Quant Lab — Daily Report</h1>
<p class="subtitle">{self.today.strftime('%A, %d %B %Y')} &nbsp;|&nbsp; Generated by IntelliSapphire</p>

<h2>Market Overview</h2>
<div class="cards">{mkt_cards}</div>

<div class="stat-row" style="margin-top:20px">
  <div class="stat"><div class="stat-val">{data['sections'].get('total_scanned', 0)}</div><div class="stat-lbl">Symbols Scanned</div></div>
  <div class="stat"><div class="stat-val" style="color:#3fb950">{data['sections'].get('total_buy', 0)}</div><div class="stat-lbl">BUY Signals</div></div>
</div>

<h2>Top BUY Signals</h2>
<table>
  <thead><tr><th>#</th><th>Symbol</th><th>Sector</th><th>Signal</th><th>Score</th><th>Conditions</th><th>Close</th><th>RSI</th></tr></thead>
  <tbody>{signal_rows if signal_rows else '<tr><td colspan="8" style="color:#8b949e;text-align:center">No BUY signals today</td></tr>'}</tbody>
</table>

<h2>Sector Performance (5-Day)</h2>
{sector_rows if sector_rows else '<p style="color:#8b949e">No sector data available.</p>'}

<div class="footer">
  NIFTY Quant Lab v1.0 &nbsp;|&nbsp; IntelliSapphire &nbsp;|&nbsp; 
  Data: NSE via Yahoo Finance &nbsp;|&nbsp; 
  Not investment advice — for educational use only.
</div>
</body>
</html>"""

            with open(html_path, "w", encoding="utf-8") as f:
                f.write(html)
            logger.info(f"HTML report: {html_path}")
            return str(html_path)

        except Exception as e:
            logger.error(f"HTML generation failed: {e}", exc_info=True)
            return None

    # ──────────────────────────────────────────────────────────
    # EXCEL REPORT
    # ──────────────────────────────────────────────────────────

    async def _generate_excel(self, data: Dict) -> Optional[str]:
        """Generate multi-sheet Excel report."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            excel_path = self.output_dir / f"report_{self.today}.xlsx"
            wb = openpyxl.Workbook()

            # ── Sheet 1: Market Overview
            ws1 = wb.active
            ws1.title = "Market Overview"
            ws1.column_dimensions["A"].width = 18
            ws1.column_dimensions["B"].width = 14
            ws1.column_dimensions["C"].width = 14
            ws1.column_dimensions["D"].width = 14
            ws1.column_dimensions["E"].width = 14
            ws1.column_dimensions["F"].width = 14

            hdr_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            hdr_font = Font(bold=True, color="FFFFFF")

            ws1["A1"] = f"NIFTY Quant Lab — Market Report: {self.today}"
            ws1["A1"].font = Font(bold=True, size=14, color="3B82F6")
            ws1.merge_cells("A1:F1")

            headers = ["Index", "Close", "Change", "Change %", "High", "Low"]
            for col, h in enumerate(headers, 1):
                cell = ws1.cell(row=3, column=col, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")

            overview = data["sections"].get("market_overview", {})
            row = 4
            for sym, d in overview.items():
                label = {"NIFTY50": "NIFTY 50", "BANKNIFTY": "BANK NIFTY", "INDIA_VIX": "India VIX"}.get(sym, sym)
                ws1.cell(row=row, column=1, value=label)
                ws1.cell(row=row, column=2, value=d["close"])
                ws1.cell(row=row, column=3, value=d["change"])
                ws1.cell(row=row, column=4, value=d["change_pct"] / 100).number_format = "0.00%"
                ws1.cell(row=row, column=5, value=d["high"])
                ws1.cell(row=row, column=6, value=d["low"])
                clr = "00C853" if d["change_pct"] >= 0 else "D50000"
                for c in [3, 4]:
                    ws1.cell(row=row, column=c).font = Font(color=clr, bold=True)
                row += 1

            # ── Sheet 2: Buy Signals
            ws2 = wb.create_sheet("Buy Signals")
            for col_w, width in zip("ABCDEFGH", [6, 14, 20, 10, 8, 12, 10, 8]):
                ws2.column_dimensions[col_w].width = width

            sig_headers = ["#", "Symbol", "Sector", "Signal", "Score", "Conditions", "Close", "RSI"]
            for col, h in enumerate(sig_headers, 1):
                cell = ws2.cell(row=1, column=col, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font

            buy_signals = data["sections"].get("buy_signals", [])
            for i, s in enumerate(buy_signals, 2):
                ws2.cell(row=i, column=1, value=i - 1)
                ws2.cell(row=i, column=2, value=s["symbol"]).font = Font(bold=True)
                ws2.cell(row=i, column=3, value=s.get("sector", ""))
                ws2.cell(row=i, column=4, value=s["signal"]).font = Font(color="00C853", bold=True)
                ws2.cell(row=i, column=5, value=s["score"])
                ws2.cell(row=i, column=6, value=f"{s['conditions']}/6")
                ws2.cell(row=i, column=7, value=s["close"])
                ws2.cell(row=i, column=8, value=s["rsi"])

            # ── Sheet 3: Sector Performance
            ws3 = wb.create_sheet("Sector Performance")
            ws3.column_dimensions["A"].width = 28
            ws3.column_dimensions["B"].width = 16
            sec_headers = ["Sector", "5-Day Return %"]
            for col, h in enumerate(sec_headers, 1):
                cell = ws3.cell(row=1, column=col, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font

            sectors = data["sections"].get("sector_performance", {})
            for i, (sec, perf) in enumerate(sectors.items(), 2):
                ws3.cell(row=i, column=1, value=sec)
                pct_cell = ws3.cell(row=i, column=2, value=perf / 100)
                pct_cell.number_format = "0.00%"
                pct_cell.font = Font(color="00C853" if perf >= 0 else "D50000", bold=True)

            wb.save(excel_path)
            logger.info(f"Excel report: {excel_path}")
            return str(excel_path)

        except ImportError:
            logger.warning("openpyxl not installed — skipping Excel report")
            return None
        except Exception as e:
            logger.error(f"Excel generation failed: {e}", exc_info=True)
            return None

    # ──────────────────────────────────────────────────────────
    # PDF (HTML → PDF conversion)
    # ──────────────────────────────────────────────────────────

    def _html_to_pdf(self, html_path: Optional[str]) -> Optional[str]:
        """Convert HTML report to PDF using weasyprint if available."""
        if not html_path or not Path(html_path).exists():
            return None
        try:
            from weasyprint import HTML
            pdf_path = html_path.replace(".html", ".pdf")
            HTML(filename=html_path).write_pdf(pdf_path)
            logger.info(f"PDF report: {pdf_path}")
            return pdf_path
        except ImportError:
            logger.debug("weasyprint not installed — no PDF output")
            return None
        except Exception as e:
            logger.warning(f"PDF generation failed: {e}")
            return None

    # ──────────────────────────────────────────────────────────
    # TELEGRAM DELIVERY
    # ──────────────────────────────────────────────────────────

    async def _send_report(self, data: Dict, paths: Dict) -> None:
        """Send report summary + files via Telegram."""
        from nifty_quant_lab.telegram.alerts import TelegramAlerter
        alerter = TelegramAlerter()
        if not alerter.is_configured:
            return

        overview = data["sections"].get("market_overview", {})
        nifty = overview.get("NIFTY50", {})
        bnk = overview.get("BANKNIFTY", {})
        vix = overview.get("INDIA_VIX", {})

        def _fmt(d: dict) -> str:
            if not d:
                return "N/A"
            arr = "▲" if d.get("change_pct", 0) >= 0 else "▼"
            return f"{d.get('close', 0):,.2f} {arr} {d.get('change_pct', 0):+.2f}%"

        msg = (
            f"📊 **NIFTY Quant Lab — EOD Report**\n"
            f"📅 {self.today.strftime('%d %b %Y (%A)')}\n\n"
            f"🔵 NIFTY:      {_fmt(nifty)}\n"
            f"🟠 BANKNIFTY:  {_fmt(bnk)}\n"
            f"⚡ India VIX:  {vix.get('close', 'N/A')}\n\n"
            f"🔍 Scanned: {data['sections'].get('total_scanned', 0)} symbols\n"
            f"✅ BUY Signals: {data['sections'].get('total_buy', 0)}\n\n"
        )

        top_buys = data["sections"].get("buy_signals", [])[:5]
        if top_buys:
            msg += "🔥 **Top 5 BUY Signals:**\n"
            for s in top_buys:
                rsi_str = f"{s['rsi']:.0f}" if s['rsi'] else 'N/A'
                msg += f"  • {s['symbol']} — Score {s['score']:.0f} | ₹{s['close']:,.2f} | RSI {rsi_str}\n"

        msg += "\n_Full report attached ↓_"
        await alerter.send_message(msg)

        # Send Excel attachment
        if paths.get("excel"):
            await alerter.send_document(
                paths["excel"],
                caption=f"📊 NIFTY Quant Lab — {self.today} Excel Report",
            )

        # Send PDF if available
        if paths.get("pdf"):
            await alerter.send_document(
                paths["pdf"],
                caption=f"📄 NIFTY Quant Lab — {self.today} PDF Report",
            )

    # ──────────────────────────────────────────────────────────
    # DB METADATA
    # ──────────────────────────────────────────────────────────

    async def _persist_report_metadata(self, data: Dict, paths: Dict) -> None:
        """Save report run metadata to daily_reports table."""
        try:
            from nifty_quant_lab.database.connection import get_async_session
            from nifty_quant_lab.database.models import DailyReport
            from nifty_quant_lab.database.upsert import mysql_upsert

            overview = data["sections"].get("market_overview", {})
            nifty = overview.get("NIFTY50", {})
            bnk = overview.get("BANKNIFTY", {})
            vix = overview.get("INDIA_VIX", {})

            async with get_async_session() as db:
                stmt = mysql_upsert(DailyReport).values(
                    report_date=self.today,
                    nifty_close=nifty.get("close"),
                    nifty_change_pct=nifty.get("change_pct"),
                    banknifty_close=bnk.get("close"),
                    banknifty_change_pct=bnk.get("change_pct"),
                    india_vix=vix.get("close"),
                    total_scanned=data["sections"].get("total_scanned", 0),
                    buy_signals=data["sections"].get("total_buy", 0),
                    html_path=paths.get("html"),
                    excel_path=paths.get("excel"),
                    pdf_path=paths.get("pdf"),
                )
                stmt = stmt.on_duplicate_key_update(
                    nifty_close=stmt.inserted.nifty_close,
                    buy_signals=stmt.inserted.buy_signals,
                    excel_path=stmt.inserted.excel_path,
                    pdf_path=stmt.inserted.pdf_path,
                )
                await db.execute(stmt)
        except Exception as e:
            logger.warning(f"Report metadata persist failed: {e}")
